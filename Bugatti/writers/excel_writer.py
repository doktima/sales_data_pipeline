# Functions for saving Excel files and formatting
import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Define yellow highlight style for error cells
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

def save_with_highlighting(df, output_file, highlight_na=True):
    """
    Save DataFrame to Excel and highlight cells containing 'NA' values.
    
    Args:
        df: DataFrame to save
        output_file: Output file path
        highlight_na: Whether to highlight NA values
    """
    try:
        # First save the DataFrame
        df.to_excel(output_file, index=False, engine="openpyxl")
        print(f"File saved to: {output_file}")
        
        if highlight_na:
            # Open the Excel file just saved
            wb = openpyxl.load_workbook(output_file)
            ws = wb.active  # Assuming first worksheet

            # Loop through rows (starting from row 2 to skip headers)
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                # Check if any cell in the row contains 'NA' (case-insensitive match)
                if any(str(cell.value).strip().upper().startswith("NA") for cell in row if cell.value is not None):
                    for cell in row:
                        cell.fill = yellow_fill

            # Save with formatting
            wb.save(output_file)
            print("Rows containing 'NA' have been highlighted in yellow.")
            
    except Exception as e:
        print(f"Failed to save file: {e}")

def create_mass_upload(combined_df, output_file):
    """
    Create a Mass Upload Excel file based on the combined data.
    
    Args:
        combined_df: DataFrame with combined data
        output_file: Output file path
    """
    try:
        # First, check if the file already exists and load it
        try:
            wb = openpyxl.load_workbook(output_file)
        except:
            # Create a new workbook if it doesn't exist
            wb = openpyxl.Workbook()
            
        ws = wb.active
        
        # Iterate through the DataFrame and populate the Excel file
        for idx, row in combined_df.iterrows():
            excel_row = idx + 2  # Start at row 2 (after headers)
            
            # Set values for each column
            ws[f"A{excel_row}"] = row["PromotionName"]
            ws[f"B{excel_row}"] = row.get("Requestor", "NA")
            ws[f"C{excel_row}"] = row.get("Start Date", "NA")
            ws[f"D{excel_row}"] = row.get("End Date", "NA")
            ws[f"E{excel_row}"] = f'=TEXT(DATE(LEFT(D{excel_row},4),MID(D{excel_row},5,2)+3,1),"YYYYMM")'
            ws[f"F{excel_row}"] = row.get("Currency", "NA")
            ws[f"G{excel_row}"] = "SAL"
            ws[f"H{excel_row}"] = ws[f"A{excel_row}"].value
            ws[f"I{excel_row}"] = row.get("Mapped Sales PGM Reason Code", "NA")
            ws[f"J{excel_row}"] = "LUMPSUM"
            ws[f"K{excel_row}"] = row.get("Customer Type", "NA")
            ws[f"L{excel_row}"] = row.get("Customer Code", "NA")
            ws[f"M{excel_row}"] = row.get("Product Type", "NA")
            ws[f"N{excel_row}"] = row.get("Model Code", "NA")
            ws[f"P{excel_row}"] = "AMT"
            ws[f"Q{excel_row}"] = row.get("Additional SOA", "NA")
            ws[f"R{excel_row}"] = row.get("Expected Sell-Out", "NA")
            ws[f"S{excel_row}"] = row.get("Expected Cost", "NA")
            ws[f"T{excel_row}"] = row.get("Apply Month", "NA")
            ws[f"U{excel_row}"] = ws[f"A{excel_row}"].value

            # Highlight rows with NA values
            highlight_yellow = any(
                str(ws.cell(row=excel_row, column=col).value).strip().upper().startswith("NA")
                for col in range(1, 22)
            )
            if highlight_yellow:
                for col in range(1, 22):
                    ws.cell(row=excel_row, column=col).fill = yellow_fill

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_len + 2

        # Save the workbook
        wb.save(output_file)
        print(f"✅ Mass Upload file created at: {output_file}")
        
    except Exception as e:
        print(f"Failed to create Mass Upload file: {e}")