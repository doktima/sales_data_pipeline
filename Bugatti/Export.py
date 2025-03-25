import os
import glob
import pandas as pd
from Headers import load_and_clean_excel, column_mapping_df
from dateutil import parser
from datetime import datetime
import calendar
from decimal import Decimal, ROUND_HALF_UP
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from dateutil.parser import parse
from team_config import TEAM_MEMBER, PATHS

# ------------------------------------------------------------------------
# 1) Constant Definitions
# ------------------------------------------------------------------------
# Sales PGM Reason Code mapping with all possible input variations
sales_pgm_reason_variations = [
    {"reason_code": "TM_R03", "variations": ["TM_R03", "DISPLAY SUPPORT REBATE"], "pgm_type": "Lumpsum", "product_type": "Division"},
    {"reason_code": "TM_R12", "variations": ["TM_R12", "NTSI", "ADDITIONAL SELL IN REBATE"], "pgm_type": "Lumpsum", "product_type": "Division"},
    {"reason_code": "TM_C01", "variations": ["TM_C01", "CO-OP", "COOP", "CO-OP AD", "CO-OP AD.", "CO-OP ADVERTISING", "COP"], "pgm_type": "Lumpsum", "product_type": "Division"},
    {"reason_code": "TM_P01", "variations": ["TM_P01", "PRICE PROTECTION"], "pgm_type": "Lumpsum", "product_type": "Model"},
    {"reason_code": "TM_Z02", "variations": ["TM_Z02", "SOA", "SELL OUT SUPPORT REBATE", "A SOA"], "pgm_type": "Lumpsum", "product_type": "Model"},
]

#Division Prefix map expect GLT and PNT
division_prefix_map = {
    "CDT": ["DB", "DF"],
    "CNT": ["GB", "GM", "GS"],
    "DFT": ["F4", "FW", "FD", "F2", "FH", "LS", "WT", "S3", "W4"]
}
# Valid Budget Allocations = Divisions
division_budget_allocations = {"GNT", "GJT", "GLT", "DFT", "CNT", "GTT", "PNT", "CDT", "PCT", "GKT"}

# AV and TV suffix logic
AV_SUFFIXES = {
    "PNT", "DGBRLLK", "CEUSCL2", ".ABEUBK", "AGBRLLK",
    ".ABEUWH", ".ABSWBK", ".ABSWWH", "AGBRLLX", "BGBRLLK",
    "EGBRLLK", "BGBRJJK", "ABEUWHF", "CGBRLLK", "AGBRLLZ",
    "CGBRLBI", "CGBRLBK", "CEUSLLK", "AEUSLLA", "AEUSLLB"
}
TV_SUFFIXES = {
    "GLT", ".AEK", "AEKQ", "AEKW", "AEKM", "AEKD"
}

# Define Folders using team configuration
source_folder = PATHS["pet_forms"] 
member_dir = PATHS["member_dir"]
base_dir = PATHS["base_dir"]  # This gives you the root project folder

# Files in team member's folder
output_file = os.path.join(member_dir, "CombinedExtractedColumns.xlsx")

# Files in root project folder
mapping_file = os.path.join(base_dir, "CustomerMapping.xlsx")
# Initialize placeholders counters
placeholder_counters = {
    'Customer Code': 0,
    'Customer Name': 0,
    'Model Code': 0,
    'Name of Promotion': 0,
    'Start Date': 0,
    'End Date': 0,
    'Expected Sell-Out': 0,
    'Additional SOA': 0
}
placeholder_value = "NA"

# ------------------------------------------------------------------------
# 2) Helper Functions
# ------------------------------------------------------------------------
# --- Single Compact Mapping Function ---
def map_all_promo_metadata(model_code, support_input):
    try:
        model_code = str(model_code).strip().upper()
        support_input = str(support_input).strip().upper()

        # Step 1: Identify Budget Allocation
        if any(model_code.endswith(suffix) for suffix in AV_SUFFIXES):
            budget_allocation = "PNT"
        elif any(model_code.endswith(suffix) for suffix in TV_SUFFIXES):
            budget_allocation = "GLT"
        elif model_code in division_budget_allocations:
            budget_allocation = model_code
        else:
            budget_allocation = "N/A"
            for division, prefixes in division_prefix_map.items():
                if any(model_code.startswith(prefix) for prefix in prefixes):
                    budget_allocation = division
                    break

        # Step 2: Determine Product Type directly from Product Code content
        product_type = "Division" if any(div in model_code for div in division_budget_allocations) else "Model"

        # Step 3: Map Reason Code and Sales PGM Type
        reason_code = "N/A"
        pgm_type = "N/A"
        
        # Special case for A SOA and similar values
        if support_input in ["A SOA", "SOA", "A-SOA"] or support_input.upper() == "SOA":
            reason_code = "TM_Z02"
            pgm_type = "Lumpsum"
        else:
            # Normal mapping logic
            for item in sales_pgm_reason_variations:
                if support_input in item["variations"]:
                    reason_code = item["reason_code"]
                    pgm_type = item["pgm_type"]
                    break

        return budget_allocation, product_type, reason_code, pgm_type

    except Exception as e:
        # Print error for debugging
        print(f"Error in mapping: {e} - model_code: {model_code}, support_input: {support_input}")
        # fallback safe default tuple
        return "N/A", "Model", "N/A", "N/A"

def parse_and_correct_date(val, is_start=True, start_reference=None):
    if pd.isna(val) or str(val).strip().upper().startswith('NA'):
        return 'NA'

    val_str = str(val).strip()
    today = datetime.today()

    # Try parsing with both formats
    try:
        parsed_dayfirst = parser.parse(val_str, dayfirst=True, yearfirst=False)
        parsed_monthfirst = parser.parse(val_str, dayfirst=False, yearfirst=False)

        # Decide which one makes more logical sense
        def date_score(dt):
            if dt.year < 2000 or dt.year > today.year + 10:  # unrealistic
                return -100
            if dt < today.replace(year=today.year - 5):  # too far in the past
                return -50
            if dt < today:
                return -10
            if abs((dt - today).days) <= 180:
                return 10
            return 0  # acceptable future

        score_dayfirst = date_score(parsed_dayfirst)
        score_monthfirst = date_score(parsed_monthfirst)

        # Select the better candidate
        parsed = parsed_dayfirst if score_dayfirst >= score_monthfirst else parsed_monthfirst

        # Validate calendar logic: (e.g., prevent 31/02 etc.)
        try:
            parsed.strftime('%Y%m%d')  # triggers ValueError for invalid calendar dates
        except ValueError:
            return 'NA'

    except Exception:
        return 'NA'

    # Final rules for start and end dates
    if is_start:
        if parsed < today:
            parsed = today
    else:
        if start_reference:
            try:
                start_dt = datetime.strptime(start_reference, '%Y%m%d')
                if parsed < start_dt:
                    parsed = start_dt
            except:
                return 'NA'

    return parsed.strftime('%Y%m%d')


def safe_round_expected_sell_out(x):
    try:
        return float(Decimal(str(x)).quantize(Decimal('1'), rounding=ROUND_HALF_UP))
    except:
        return 0  # or np.nan or any fallback value you want

def parse_flexible_date(val):
    if pd.isna(val):
        return pd.NaT
    try:
        return parser.parse(str(val), dayfirst=False, yearfirst=False)
    except:
        try:
            return parser.parse(str(val), dayfirst=True, yearfirst=False)
        except:
            return pd.NaT

def get_apply_months_and_days(start, end):
    months = []
    try:
        start_dt = datetime.strptime(start, '%Y%m%d')
        end_dt = datetime.strptime(end, '%Y%m%d')
        current = start_dt.replace(day=1)
        while current <= end_dt:
            month_str = current.strftime('%Y%m')
            first_day = current
            last_day = datetime(current.year, current.month, calendar.monthrange(current.year, current.month)[1])
            range_start = max(start_dt, first_day)
            range_end = min(end_dt, last_day)
            days_in_month = (range_end - range_start).days + 1
            months.append((month_str, days_in_month))
            # Move to next month
            if current.month == 12:
                current = current.replace(year=current.year + 1, month=1)
            else:
                current = current.replace(month=current.month + 1)
    except:
        pass
    return months

def is_likely_customer_code(val):
    valid_code_prefixes = ('IE', 'GB', '50', 'OB', 'JE', 'GG', '55', 'OB', '11', 'JE', '18', 'SE', 'HE', 'RA', '74', '13', '10')
    exceptions = {'HETIER1', 'HETIER2', 'HEBNO', 'SEVENOAKS_AWE', 'HEKEYINDY', 'RADIUS_CIH', 'OBSIDIAN', '50380042-S'}
    
    if not isinstance(val, str):
        return False
    return val.upper().startswith(valid_code_prefixes) or val.upper() in exceptions

def is_likely_customer_name(val):
    if not isinstance(val, str):
        return False
    return not is_likely_customer_code(val)

def detect_errors(row):
    errors = []

    if row.get('Customer Type', '') == 'NA':
        errors.append('Missing Customer Type')
    if row.get('Requestor', '') == 'NA':
        errors.append('Missing Requestor')
    if row.get('Currency', '') == 'NA':
        errors.append('Missing Currency')
    if str(row.get('Start Date', '')).startswith('N/A'):
        errors.append('Invalid Start Date')
    if str(row.get('End Date', '')).startswith('N/A'):
        errors.append('Invalid End Date')
    if pd.isna(row.get('Model Code')) or str(row.get('Model Code')).strip() in ['NA', '']:
        errors.append('Missing Model Code')
    if pd.isna(row.get('Additional SOA')):
        errors.append('Missing Additional SOA')
    if pd.isna(row.get('Expected Sell-Out')):
        errors.append('Missing Expected Sell-Out')

    return ', '.join(errors) if errors else ''

# ------------------------------------------------------------------------
# 3) Main Processing
# ------------------------------------------------------------------------
# Initialize Final Combined DataFrame
combined_df = pd.DataFrame()
df_mapping = pd.read_excel(mapping_file)

# Loop Through Excel Files
excel_files = glob.glob(os.path.join(source_folder, "*.xlsx"))
if not excel_files:
    print("❌ No Excel files found in source folder.")
else:
    print(f"📂 Found {len(excel_files)} files. Extracting specific columns...")

    for file_path in excel_files:
        print(f"\n📄 Processing: {os.path.basename(file_path)}")

        # ✅ Step 1: Load clean data
        cleaned_df = load_and_clean_excel(file_path)
        if cleaned_df is None:
            print("⚠️ Skipping due to read/clean error.")
            continue

        # ✅ Clean headers to avoid hidden formatting mismatches
        cleaned_df.columns = cleaned_df.columns.astype(str).str.strip().str.replace(r'\s+', ' ', regex=True)

        print("Columns after cleaning:", cleaned_df.columns.tolist())
        cleaned_df = cleaned_df.loc[:, ~cleaned_df.columns.duplicated()]

        print(f"➡️ Loaded rows: {len(cleaned_df)} from {file_path}")

        # ✅ Step 2: Extract only required columns
        required_columns = column_mapping_df['Standard Column'].tolist()
        extracted_df = cleaned_df.reindex(columns=required_columns)

        # ✅ Step 2.1: Fill missing or empty columns with dynamic placeholders 
        for col in required_columns:
            if col not in extracted_df.columns or extracted_df[col].isnull().all():
                count = placeholder_counters.get(col, 0)
                placeholder_value = "NA"
                placeholder_counters[col] = count + 1

                if col in ['Expected Sell-Out', 'Additional SOA']:
                    extracted_df[col] = 0
                elif col in ['Start Date', 'End Date']:
                    extracted_df[col] = '19000101'
                else:
                    extracted_df[col] = placeholder_value

        # ✅ Step 2.2: Handle missing Type of Support
        if 'Type of Support' not in extracted_df.columns:
            # If column doesn't exist at all, create it
            extracted_df['Type of Support'] = 'A SOA'
            print("📝 Added missing 'Type of Support' column with default 'A SOA' value")
        else:
            # Fix missing/NA values in Type of Support
            extracted_df['Type of Support'] = extracted_df['Type of Support'].fillna('A SOA')
            
            # Find rows with placeholder values like NA, empty string, etc.
            missing_mask = (
                (extracted_df['Type of Support'].astype(str).str.strip() == '') | 
                (extracted_df['Type of Support'].astype(str).str.upper().isin(['NA', 'N/A', 'NONE', '-', 'NULL']))
            )
            
            if missing_mask.sum() > 0:
                extracted_df.loc[missing_mask, 'Type of Support'] = 'A SOA'
                print(f"📝 Set 'A SOA' for {missing_mask.sum()} rows with missing Type of Support")

        # Step 2.3: Auto-fix swapped Customer Name & Customer Code if needed
        mask_swapped = extracted_df.apply(
            lambda row: is_likely_customer_code(row['Customer Name']) and is_likely_customer_name(row['Customer Code']),
            axis=1
        )
        extracted_df.loc[mask_swapped, ['Customer Name', 'Customer Code']] = extracted_df.loc[mask_swapped, ['Customer Code', 'Customer Name']].values

        # Round Additional SOA
        extracted_df['Additional SOA'] = pd.to_numeric(extracted_df['Additional SOA'], errors='coerce').round(2)

        # ✅ Step 3: Normalize Dates
        extracted_df['Start Date'] = extracted_df['Start Date'].apply(lambda x: parse_and_correct_date(x, is_start=True))
        extracted_df['End Date'] = extracted_df.apply(
            lambda row: parse_and_correct_date(row['End Date'], is_start=False, start_reference=row['Start Date']),
            axis=1
        )
        # ✅ Step 4: Drop invalid rows
        extracted_df = extracted_df.dropna(subset=['Model Code', 'Additional SOA', 'Start Date', 'End Date'])
        extracted_df = extracted_df[extracted_df.notna().sum(axis=1) >= 3]

        # ✅ Step 5: Convert Expected Sell-Out to numeric and round
        extracted_df['Expected Sell-Out'] = pd.to_numeric(extracted_df['Expected Sell-Out'], errors='coerce').fillna(0)
        extracted_df['Expected Sell-Out'] = extracted_df['Expected Sell-Out'].apply(lambda x: float(Decimal(x).quantize(Decimal('1'), rounding=ROUND_HALF_UP)))

        # 🔒 Preserve Row Order
        extracted_df['Original Row Index'] = range(len(extracted_df))
        extracted_df['Source File'] = os.path.basename(file_path)
        
        # ✅ Extract and clean WBW TV MODEL column only
        wbw_candidates = [col for col in cleaned_df.columns if 'WBW' in col.upper() and 'MODEL' in col.upper()]
        if wbw_candidates:
            wbw_col = wbw_candidates[0]
            print(f"🔍 Found WBW column: '{wbw_col}'")
            
            # Clean and standardize WBW values
            wbw_cleaned = cleaned_df[wbw_col].astype(str).str.strip().str.upper()
            
            # Replace obvious placeholder values
            invalid_values = ['NA', 'NA1', 'NA2', 'NA3', 'NO TV MODEL', '', 'NONE', 'N/A', 'NO', 'NULL']
            for invalid in invalid_values:
                wbw_cleaned = wbw_cleaned.replace(invalid, 'NO TV MODEL')
            
            extracted_df['WBW TV MODEL'] = wbw_cleaned
            
            # Set Is WBW flag based on cleaned values
            extracted_df['Is WBW'] = "NO"
            extracted_df.loc[(extracted_df['WBW TV MODEL'] != 'NO TV MODEL') & 
                            (extracted_df['WBW TV MODEL'].str.len() > 3), 'Is WBW'] = "YES"
            
            # Print diagnostic info
            wbw_count = (extracted_df['Is WBW'] == "YES").sum()
            print(f"🔍 WBW rows identified in this file: {wbw_count}")
            if wbw_count > 0:
                print(f"🔍 Sample WBW values: {extracted_df.loc[extracted_df['Is WBW'] == 'YES', 'WBW TV MODEL'].unique()[:3]}")
        else:
            # No WBW column found
            extracted_df['WBW TV MODEL'] = 'NO TV MODEL'
            extracted_df['Is WBW'] = "NO"

        # ✅ Update Name of Promotion for WBW rows before grouping
        if 'Is WBW' in extracted_df.columns:
            wbw_mask = extracted_df['Is WBW'] == 'YES'
            if wbw_mask.sum() > 0:
                # Concatenate Name of Promotion + Model Code for WBW rows
                extracted_df.loc[wbw_mask, 'Name of Promotion'] = (
                    extracted_df.loc[wbw_mask, 'Name of Promotion'].fillna('').astype(str).str.strip() + " " +
                    extracted_df.loc[wbw_mask, 'Model Code'].fillna('').astype(str).str.strip()
                ).str.strip()
                print(f"🔍 Updated {wbw_mask.sum()} promotion names for WBW rows")

        # 🔒 Preserve Row Order
        extracted_df['Original Row Index'] = range(len(extracted_df))
        extracted_df['Source File'] = os.path.basename(file_path)

        # 🔄 Define grouping columns
        group_cols = ['Customer Name', 'Customer Code', 'Model Code', 'Start Date', 'End Date', 'Additional SOA', 'Source File', 'Name of Promotion', 'Is WBW']

        # Make sure all group_cols exist
        for col in group_cols:
            if col not in extracted_df.columns:
                print(f"⚠️ Missing column for grouping: {col}")
                extracted_df[col] = 'NA'  # Add placeholder

        # Define aggregation functions
        agg_dict = {col: 'first' for col in extracted_df.columns if col not in group_cols}
        agg_dict['Expected Sell-Out'] = 'sum'  # Sum quantities

        # Print before grouping stats
        print(f"➡️ Before grouping: {len(extracted_df)} rows, {extracted_df['Expected Sell-Out'].sum()} units")

        # Perform the grouping
        grouped_df = extracted_df.groupby(group_cols, as_index=False).agg(agg_dict)

        # Print after grouping stats
        print(f"➡️ After grouping: {len(grouped_df)} rows, {grouped_df['Expected Sell-Out'].sum()} units")

        #  Expand by Apply Month & Distribute Quantity
        expanded_rows = []
        for _, row in grouped_df.iterrows():
            start, end = row['Start Date'], row['End Date']
            try:
                months = get_apply_months_and_days(start, end)
                if not months:
                    row['Apply Month'] = 'NA'
                    expanded_rows.append(row)
                    continue

                total_days = sum(d for _, d in months)
                qty = row['Expected Sell-Out']

                if qty == 1:
                    dist_qty = [1] * len(months)
                elif qty == 2:
                    dist_qty = [1, 1] if len(months) == 2 else [2] if len(months) == 1 else [1] * len(months)
                elif qty == 3:
                    dist_qty = [3] if len(months) == 1 else [2, 1] if len(months) == 2 else [1] * len(months)
                else:
                    dist_qty = []
                    remaining = qty
                    for i, (_, days) in enumerate(months[:-1]):
                        part = round(qty * days / total_days)
                        dist_qty.append(part)
                        remaining -= part
                    dist_qty.append(max(0, remaining))

                for (month, _), qty_month in zip(months, dist_qty):
                    new_row = row.copy()
                    new_row['Apply Month'] = month
                    new_row['Expected Sell-Out'] = qty_month
                    new_row['Errors in Combined Extract'] = ''
                    expanded_rows.append(new_row)

            except Exception as e:
                row['Apply Month'] = 'NA'
                row['Errors in Combined Extract'] = f"Expansion error: {e}"
                expanded_rows.append(row)

        if expanded_rows:
            expanded_df = pd.DataFrame(expanded_rows)
            expanded_df = expanded_df.sort_values(by='Original Row Index').reset_index(drop=True)
            combined_df = pd.concat([combined_df, expanded_df], ignore_index=True)
        else:
            print(f"⚠️ No valid rows found for expansion in: {os.path.basename(file_path)}")
# ------------------------------------------------------------------------
# 4) Post-Processing
# ------------------------------------------------------------------------
# Force fix any remaining NA Type of Support values
if 'Type of Support' in combined_df.columns:
    na_support_mask = (
        pd.isna(combined_df['Type of Support']) | 
        (combined_df['Type of Support'].astype(str).str.strip() == '') |
        (combined_df['Type of Support'].astype(str).str.upper().isin(['NA', 'N/A', 'NONE', '-']))
    )
    
    na_count = na_support_mask.sum()
    if na_count > 0:
        combined_df.loc[na_support_mask, 'Type of Support'] = 'A SOA'
        print(f"📝 Fixed {na_count} rows with NA Type of Support values in final processing")

# ✅ Calculate Total SOA
combined_df['Expected Cost'] = (combined_df['Additional SOA'] * combined_df['Expected Sell-Out']).round(2)

# ✅ Normalize casing for Customer Code in both DataFrames
combined_df['Customer Code'] = combined_df['Customer Code'].astype(str).str.strip().str.upper()
df_mapping['Customer Code'] = df_mapping['Customer Code'].astype(str).str.strip().str.upper()

# ✅ (Optional: still standardize Budget Allocation column if you need it elsewhere)
if 'Budget Allocation' in df_mapping.columns:
    df_mapping['Budget Allocation'] = df_mapping['Budget Allocation'].astype(str).str.strip().str.upper()

# ✅ Drop duplicates from mapping to keep only the first match per Customer Code
df_mapping = df_mapping.drop_duplicates(subset='Customer Code', keep='first')

# ✅ Merge with mapping (only on Customer Code)
combined_df = combined_df.merge(
    df_mapping[['Customer Code', 'Customer Type', 'Requestor', 'Currency']],
    on='Customer Code',
    how='left'
)

# ✅ Fill missing values if any field is not mapped
combined_df[['Customer Type', 'Requestor', 'Currency']] = combined_df[['Customer Type', 'Requestor', 'Currency']].fillna('NA')

# ✅ Recalculate Budget Allocation after mapping if needed (optional)
combined_df['Budget Allocation'] = combined_df['Model Code'].apply(lambda x: map_all_promo_metadata(x, "")[0])

# ✅ Apply error detection fields based on mapping logic
combined_df[['Budget Allocation', 'Product Type', 'Mapped Sales PGM Reason Code', 'Sales PGM Type']] = combined_df.apply(
    lambda row: pd.Series(map_all_promo_metadata(row['Model Code'], row.get('Type of Support', ''))),
    axis=1
)

# ------------------------------------------------------------------------
# 7) Save Final Combined File
# ------------------------------------------------------------------------
try:
    combined_df.to_excel(output_file, index=False, engine="openpyxl")
    print(f"\n✅ Final file with distributed sell-out saved to: {output_file}")
    # Yellow fill style
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Open the Excel file just saved
    wb = load_workbook(output_file)
    ws = wb.active  # Assuming first worksheet

    # Loop through rows (starting from row 2 to skip headers)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        # Check if any cell in the row contains 'NA' (case-insensitive match)
        if any(str(cell.value).strip().upper().startswith("NA") for cell in row if cell.value is not None):
            for cell in row:
                cell.fill = yellow_fill

    # Save with formatting
    wb.save(output_file)
    print("🎨 Rows containing 'NA' have been highlighted in yellow.")
except Exception as e:
    print(f"❌ Failed to save final file: {e}")