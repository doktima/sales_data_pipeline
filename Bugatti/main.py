import os
import sys
import glob
import pandas as pd
from datetime import datetime

# Add the project root to Python path
script_dir = os.path.dirname(os.path.abspath(__file__))
if script_dir not in sys.path:
    sys.path.insert(0, script_dir)

from config.paths import PATHS, TEAM_MEMBER
from etl.loader import load_and_clean_excel
from etl.parser import parse_and_correct_date, is_likely_customer_code, is_likely_customer_name, standardize_customer_code
from etl.mapping import map_all_promo_metadata, classify_model_code
from etl.grouping import group_similar_rows, distribute_quantities_by_month
from etl.validation import detect_errors
from writers.excel_writer import save_with_highlighting, create_mass_upload
from writers.promo_naming import build_name_of_promotion

def process_pet_forms():
    """
    Main function to process PET forms, extract data, and create output files.
    """
    print(f"Running script for team member: {TEAM_MEMBER}")
    print(f"Source folder: {PATHS['pet_forms']}")
    print(f"Output folder: {PATHS['uploads']}")
    
    # Make sure output directories exist
    os.makedirs(PATHS['uploads'], exist_ok=True)
    
    # Define output files
    combined_file = os.path.join(PATHS['member_dir'], "CombinedExtractedColumns.xlsx")
    mass_upload_file = os.path.join(PATHS['uploads'], "MassUpload.xlsx")
    # -------------------------------------------------------------------
    # CLEAN PREVIOUS FILES
    # -------------------------------------------------------------------

    # 1. Remove CombinedExtractedColumns file completely
    if os.path.exists(combined_file):
        os.remove(combined_file)
        print("Removed old CombinedExtractedColumns.xlsx")

    # 2. Clean all MassUpload content except header row
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    if os.path.exists(mass_upload_file):
        try:
            wb = load_workbook(mass_upload_file)
            ws = wb.active
            ws.delete_rows(2, ws.max_row)  # remove everything below the first row
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.fill = PatternFill()  # clear fill color
            wb.save(mass_upload_file)
            print("Cleaned MassUpload.xlsx (kept only header row)")
        except Exception as e:
            print(f"Could not clean MassUpload.xlsx: {e}")

    
    # Get customer mapping data
    mapping_file = os.path.join(PATHS['base_dir'], "CustomerMapping.xlsx")
    try:
        df_mapping = pd.read_excel(mapping_file)
        df_mapping['Customer Code'] = df_mapping['Customer Code'].astype(str).str.strip().str.upper()
        # Drop duplicates from mapping to keep only the first match
        df_mapping = df_mapping.drop_duplicates(subset='Customer Code', keep='first')
    except Exception as e:
        print(f"Error loading customer mapping: {e}")
        df_mapping = pd.DataFrame(columns=['Customer Code', 'Customer Type', 'Requestor', 'Currency'])
    
    # Find Excel files
    excel_files = glob.glob(os.path.join(PATHS['pet_forms'], "*.xlsx"))
    if not excel_files:
        print("No Excel files found in source folder.")
        return
    
    print(f"Found {len(excel_files)} files. Processing...")
    
    # Initialize combined DataFrame
    combined_df = pd.DataFrame()
    
    # Process each Excel file
    for file_path in excel_files:
        print(f"Processing: {os.path.basename(file_path)}")
        
        # Step 1: Load and clean Excel
        cleaned_df = load_and_clean_excel(file_path)
        if cleaned_df is None:
            print("Skipping due to read/clean error.")
            continue
        
        # Clean headers to avoid hidden formatting mismatches
        cleaned_df.columns = cleaned_df.columns.astype(str).str.strip().str.replace(r'\s+', ' ', regex=True)
        
        # Remove duplicate columns if any
        cleaned_df = cleaned_df.loc[:, ~cleaned_df.columns.duplicated()]
        
        print(f"➡️ Loaded rows: {len(cleaned_df)} from {file_path}")
        
        # Step 2: Get required columns (based on column mapping)
        required_columns = [
            'Customer Code', 'Customer Name', 'Model Code', 'Type of Support', 
            'Additional SOA', 'Expected Sell-Out', 'Start Date', 'End Date', 
            'Expected Cost', 'Name of Promotion'
        ]
        
        # Create a copy for extraction
        extracted_df = cleaned_df.reindex(columns=required_columns)
        
        # Fill missing columns with default values
        for col in required_columns:
            if col not in extracted_df.columns or extracted_df[col].isnull().all():
                if col in ['Expected Sell-Out', 'Additional SOA']:
                    extracted_df[col] = 0
                elif col in ['Start Date', 'End Date']:
                    extracted_df[col] = '19000101'
                else:
                    extracted_df[col] = "NA"
        
        # Handle missing Type of Support
        if 'Type of Support' not in extracted_df.columns:
            extracted_df['Type of Support'] = 'A SOA'
        else:
            extracted_df['Type of Support'] = extracted_df['Type of Support'].astype(str)
            extracted_df['Type of Support'] = extracted_df['Type of Support'].fillna('A SOA')
            
            # Fix missing/placeholder values
            missing_mask = (
                (extracted_df['Type of Support'].str.strip() == '') | 
                (extracted_df['Type of Support'].str.upper().isin(['NA', 'N/A', 'NONE', '-', 'NULL'])) |
                (extracted_df['Type of Support'].str.isnumeric()) |
                (extracted_df['Type of Support'].str.replace('.', '', regex=False).str.isnumeric())
            )
            
            if missing_mask.sum() > 0:
                extracted_df.loc[missing_mask, 'Type of Support'] = 'A SOA'
        
        # Standardize customer codes first (before swap detection)
        extracted_df['Customer Code'] = extracted_df['Customer Code'].astype(str).fillna('NA')
        extracted_df['Customer Name'] = extracted_df['Customer Name'].astype(str).fillna('NA')
        
        # Standardize Customer Codes - apply proper case formatting
        extracted_df['Customer Code'] = extracted_df['Customer Code'].apply(standardize_customer_code)
        
        # Auto-fix swapped Customer Name & Customer Code if needed
        mask_swapped = extracted_df.apply(
            lambda row: is_likely_customer_code(row['Customer Name']) and is_likely_customer_name(row['Customer Code']),
            axis=1
        )
        
        if mask_swapped.sum() > 0:
            # Store the swapped values temporarily 
            temp_codes = extracted_df.loc[mask_swapped, 'Customer Name'].apply(standardize_customer_code)
            temp_names = extracted_df.loc[mask_swapped, 'Customer Code']
            
            # Apply the swap
            extracted_df.loc[mask_swapped, 'Customer Code'] = temp_codes
            extracted_df.loc[mask_swapped, 'Customer Name'] = temp_names
            
            print(f"Fixed {mask_swapped.sum()} rows with swapped customer code/name")
        
        # Ensure all customer codes are properly standardized after potential swaps
        extracted_df['Customer Code'] = extracted_df['Customer Code'].apply(standardize_customer_code)
        
        # Round Additional SOA to 2 decimal places
        extracted_df['Additional SOA'] = pd.to_numeric(extracted_df['Additional SOA'], errors='coerce').round(2)
        
        # Normalize dates
        extracted_df['Start Date'] = extracted_df['Start Date'].apply(lambda x: parse_and_correct_date(x, is_start=True))
        extracted_df['End Date'] = extracted_df.apply(
            lambda row: parse_and_correct_date(row['End Date'], is_start=False, start_reference=row['Start Date']),
            axis=1
        )
        
        # Convert Expected Sell-Out to numeric and round
        extracted_df['Expected Sell-Out'] = pd.to_numeric(extracted_df['Expected Sell-Out'], errors='coerce').fillna(0)
        extracted_df['Expected Sell-Out'] = extracted_df['Expected Sell-Out'].round(0)
        
        # Preserve row order and source file info
        extracted_df['Original Row Index'] = range(len(extracted_df))
        extracted_df['Source File'] = os.path.basename(file_path)
        
        # Extract WBW TV MODEL if available
        wbw_candidates = [col for col in cleaned_df.columns if 'WBW' in col.upper() and 'MODEL' in col.upper()]
        if wbw_candidates:
            wbw_col = wbw_candidates[0]
            print(f"Found WBW column: '{wbw_col}'")
            
            # Clean and standardize WBW values
            wbw_cleaned = cleaned_df[wbw_col].astype(str).str.strip().str.upper()
            
            # Replace obvious placeholder values
            invalid_values = ['NA', 'NA1', 'NA2', 'NA3', 'NO TV MODEL', '', 'NONE', 'N/A', 'NO', 'NULL']
            for invalid in invalid_values:
                wbw_cleaned = wbw_cleaned.replace(invalid, 'NO TV MODEL')
            
            extracted_df['WBW TV MODEL'] = wbw_cleaned
            
            # Set Is WBW flag based on cleaned values
            extracted_df['Is WBW'] = "NO"
            wbw_mask = (extracted_df['WBW TV MODEL'] != 'NO TV MODEL') & (extracted_df['WBW TV MODEL'].str.len() > 3)
            extracted_df.loc[wbw_mask, 'Is WBW'] = "YES"
            
            # Update Name of Promotion for WBW rows
            if wbw_mask.sum() > 0:
                extracted_df.loc[wbw_mask, 'Name of Promotion'] = (
                    extracted_df.loc[wbw_mask, 'Name of Promotion'].fillna('').astype(str).str.strip() + " " +
                    extracted_df.loc[wbw_mask, 'Model Code'].fillna('').astype(str).str.strip()
                ).str.strip()
                print(f"Updated {wbw_mask.sum()} promotion names for WBW rows")
        else:
            # No WBW column found
            extracted_df['WBW TV MODEL'] = 'NO TV MODEL'
            extracted_df['Is WBW'] = "NO"
        
        # Group similar rows
        grouped_df = group_similar_rows(extracted_df)
        print(f"➡️ After grouping: {len(grouped_df)} rows, {grouped_df['Expected Sell-Out'].sum()} units")
        
        # Expand by Apply Month & Distribute Quantity
        expanded_df = distribute_quantities_by_month(grouped_df)
        
        # Add to combined data
        if not expanded_df.empty:
            combined_df = pd.concat([combined_df, expanded_df], ignore_index=True)
        else:
            print(f"No valid rows found for expansion in: {os.path.basename(file_path)}")
    
    # Post-processing
    if not combined_df.empty:
        # Force fix any remaining NA Type of Support values
        if 'Type of Support' in combined_df.columns:
            na_support_mask = (
                pd.isna(combined_df['Type of Support']) | 
                (combined_df['Type of Support'].astype(str).str.strip() == '') |
                (combined_df['Type of Support'].astype(str).str.upper().isin(['NA', 'N/A', 'NONE', '-']))
            )
            
            na_count = na_support_mask.sum()
            if na_count > 0:
                combined_df.loc[na_support_mask, 'Type of Support'] = 'A SOA'
                print(f"📝 Fixed {na_count} rows with NA Type of Support values in final processing")
        
        # Final check to ensure all customer codes are standardized
        combined_df['Customer Code'] = combined_df['Customer Code'].apply(standardize_customer_code)
        
        # Calculate Total SOA (Expected Cost)
        combined_df['Expected Cost'] = (combined_df['Additional SOA'] * combined_df['Expected Sell-Out']).round(2)
        
        # Merge with customer mapping
        combined_df = combined_df.merge(
            df_mapping[['Customer Code', 'Customer Type', 'Requestor', 'Currency']],
            on='Customer Code',
            how='left'
        )
        
        # Fill missing mapped values
        combined_df[['Customer Type', 'Requestor', 'Currency']] = combined_df[['Customer Type', 'Requestor', 'Currency']].fillna('NA')
        
        # Apply mapping logic
        combined_df[['Budget Allocation', 'Product Type', 'Mapped Sales PGM Reason Code', 'Sales PGM Type']] = combined_df.apply(
            lambda row: pd.Series(map_all_promo_metadata(row['Model Code'], row.get('Type of Support', ''))),
            axis=1
        )
        
        # Classify model codes
        combined_df['Segment'] = combined_df['Model Code'].apply(classify_model_code)
        
        # Build promotion names
        combined_df['PromotionName'] = combined_df.apply(build_name_of_promotion, axis=1)
        
        # Save combined file
        save_with_highlighting(combined_df, combined_file)
        
        # Create MassUpload file
        create_mass_upload(combined_df, mass_upload_file)
        
        print("Processing completed successfully.")
    else:
        print("No valid data found for processing.")

if __name__ == "__main__":
    process_pet_forms()