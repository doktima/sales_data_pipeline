import os
import openpyxl
import pandas as pd
import re
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
from team_config import TEAM_MEMBER, PATHS

# File paths using team configuration
base_dir = PATHS["base_dir"]
member_dir = PATHS["member_dir"]
uploads_dir = PATHS["uploads"]

# Files specific to this script - CombinedExtractedColumns.xlsx is in member_dir (Tima folder)
COMBINED_FILE = os.path.join(member_dir, "CombinedExtractedColumns.xlsx")
MASS_UPLOAD_FILE = os.path.join(uploads_dir, "MassUpload.xlsx")

ABBREVIATIONS = {"AV", "TV", "SOA", "UK", "EE", "QNED", "OLED", "HDR", "UHD", "AI", "LG", "FOC", "WBW"}
REMOVE_WORDS = {"CIH", "EXRTIS"}

# AV and TV suffix sets
AV_SUFFIXES = {
    "PNT", "DGBRLLK", "CEUSCL2", ".ABEUBK", "AGBRLLK",
    ".ABEUWH", ".ABSWBK", ".ABSWWH", "AGBRLLX", "BGBRLLK",
    "EGBRLLK", "BGBRJJK", "ABEUWHF", "CGBRLLK", "AGBRLLZ",
    "CGBRLBI", "CGBRLBK", "CEUSLLK", "AEUSLLA", "AEUSLLB"
}
TV_SUFFIXES = {
    "GLT", ".AEK", "AEKQ", "AEKW", "AEKM", "AEKD"
}

HS_CODES = {"CDT", "CNT", "DFT"}

yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

def classify_model_code(model_code):
    if not isinstance(model_code, str):
        return "UNKNOWN"
    if any(model_code.endswith(sfx) for sfx in AV_SUFFIXES):
        return "AV"
    if any(model_code.endswith(sfx) for sfx in TV_SUFFIXES):
        return "TV"
    return "UNKNOWN"

def format_title_case(text):
    words = re.split(r'(\s+)', text)
    formatted_words = []
    for word in words:
        clean_word = word.strip().upper()
        if clean_word in REMOVE_WORDS:
            continue
        if clean_word in ABBREVIATIONS:
            formatted_words.append(word.upper())
        else:
            formatted_words.append(word.capitalize())
    result = ''.join(formatted_words)
    result = re.sub(r'\s+', ' ', result).strip()
    return result

def safe_get(row, colname):
    value = row.get(colname, 'NA')
    if pd.isna(value) or str(value).strip() == '':
        return 'NA'
    return value

def build_name_of_promotion(row):
    customer = safe_get(row, "Customer Name")
    segment = safe_get(row, "Segment")
    promo = safe_get(row, "Name of Promotion")
    start = safe_get(row, "Start Date")
    end = safe_get(row, "End Date")
    budget_alloc = safe_get(row, "Budget Allocation")
    support_type = safe_get(row, "Type of Support")

    promo_clean = " ".join(word for word in str(promo).split() if word.upper() not in REMOVE_WORDS)

    if budget_alloc in HS_CODES:
        if "PRM" not in str(promo).upper():
            full_promo = f"HS - PET - {budget_alloc} - {promo_clean} - {support_type} - {customer} - {start} TO {end}"
        else:
            full_promo = f"HS - PRM - {budget_alloc} - {promo_clean} - {support_type} - {customer} - {start} TO {end}"
    else:
        base = f"{customer} ({segment})"
        full_promo = f"{base} {promo_clean} PET {support_type} {start} TO {end}"

    full_promo = re.sub(r'\s+', ' ', full_promo).strip().upper()
    return full_promo

def main():
    df_combined = pd.read_excel(COMBINED_FILE)

    df_combined["Start Date"] = df_combined["Start Date"].astype(str).str.strip()
    df_combined["End Date"] = df_combined["End Date"].astype(str).str.strip()
    df_combined["Segment"] = df_combined["Model Code"].apply(classify_model_code)
    df_combined["PromotionName"] = df_combined.apply(build_name_of_promotion, axis=1)

    wb = openpyxl.load_workbook(MASS_UPLOAD_FILE)
    ws = wb.active

    for idx, row in df_combined.iterrows():
        excel_row = idx + 2
        start_str = safe_get(row, "Start Date")
        end_str = safe_get(row, "End Date")
        ws[f"A{excel_row}"] = row["PromotionName"]
        ws[f"B{excel_row}"] = safe_get(row, "Requestor")
        ws[f"C{excel_row}"] = start_str
        ws[f"D{excel_row}"] = end_str
        ws[f"E{excel_row}"] = f'=TEXT(DATE(LEFT(D{excel_row},4),MID(D{excel_row},5,2)+3,1),"YYYYMM")'
        ws[f"F{excel_row}"] = safe_get(row, "Currency")
        ws[f"G{excel_row}"] = "SAL"
        ws[f"H{excel_row}"] = ws[f"A{excel_row}"].value
        ws[f"I{excel_row}"] = safe_get(row, "Mapped Sales PGM Reason Code")
        ws[f"J{excel_row}"] = "LUMPSUM"
        ws[f"K{excel_row}"] = safe_get(row, "Customer Type")
        ws[f"L{excel_row}"] = safe_get(row, "Customer Code")
        ws[f"M{excel_row}"] = safe_get(row, "Product Type")
        ws[f"N{excel_row}"] = safe_get(row, "Model Code")
        ws[f"P{excel_row}"] = "AMT"
        ws[f"Q{excel_row}"] = safe_get(row, "Additional SOA")
        ws[f"R{excel_row}"] = safe_get(row, "Expected Sell-Out")
        ws[f"S{excel_row}"] = safe_get(row, "Expected Cost")
        ws[f"T{excel_row}"] = safe_get(row, "Apply Month")
        ws[f"U{excel_row}"] = ws[f"A{excel_row}"].value

        highlight_yellow = any(
            str(ws.cell(row=excel_row, column=col).value).strip().upper().startswith("NA")
            for col in range(1, 22)
        )
        if highlight_yellow:
            for col in range(1, 22):
                ws.cell(row=excel_row, column=col).fill = yellow_fill

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(MASS_UPLOAD_FILE)
    print("MassUpload file updated and saved successfully.")

if __name__ == "__main__":
    main()
